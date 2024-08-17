from openpyxl import load_workbook
from datetime import datetime

# Abrir o arquivo Excel existente
workbook = load_workbook('files/billing.xlsx')
pg_bill = workbook['Sheet1']

# Iterar sobre as linhas a partir da segunda linha
for row in pg_bill.iter_rows(min_row=2):
    vencimento_str = row[2].value  # Lê a coluna "Vencimento" (índice 2)

    try:
        datetime.strptime(vencimento_str, '%d/%m/%Y')
    except ValueError:
        print(f"Data inválida na linha {row[0].row}: {vencimento_str}")