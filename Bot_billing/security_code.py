import os
import webbrowser
from datetime import datetime
from time import sleep
from urllib.parse import quote

from openpyxl import load_workbook
import pyautogui

# ... resto do seu código ...

#webbrowser.open('https://web.whatsapp.com/')
#print('abri o navegador inicial, vou esperar 20s')
#sleep(15)

print('abrindo lendo arquivo')
# Abrir o arquivo Excel existente
workbook = load_workbook('files/billing.xlsx')
pg_bill = workbook['Sheet1']

date_now = datetime.now()

for row in pg_bill.iter_rows(min_row=2, max_row=22):
 company = row[0].value
phone = row[1].value
bill_str = row[2].value  # Leia a data como string
bill = datetime.strptime(bill_str, '%d/%m/%Y')  # Converta para datetime

if bill < date_now:
    payment = 'venceu'
else:
    payment = 'vai vencer'

msg = f'Olá {company} seu boleto {payment} no dia {bill.strftime('%d/%m/%Y')}. Favor pagar no link https://www.seusistemapagamento.com'

try:
    print('vou abrir navegador')
    link_msg_whats = f'https://web.whatsapp.com/send?phone={phone}&text={quote(msg)}'
    webbrowser.open(link_msg_whats)
    print('acabei de abrir agora vou esperar 20s')
    sleep(20)
    print('vou localizar a seta')
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    seta = pyautogui.locateCenterOnScreen('files/arrow.png')
    print('localizei a seta vou esperar 5s')
    sleep(5)
    print('vou clicar')
    pyautogui.click(seta[0], seta[1])
    print('cliquei vou esperar 5s')
    sleep(5)
    print('esperei agora vou fechar')
    pyautogui.hotkey('ctrl', 'w')
    print('fechei agora vou esperar 2s')
    sleep(2)
    print('fechei agora vou repetir a ação')

except:
    print(f'Não foi possível enviar mensagem para {company}')
    with open('files/erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
        arquivo.write(f'{company},{phone}{os.linesep}')

