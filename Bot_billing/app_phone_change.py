from openpyxl import load_workbook

# Abrir o arquivo Excel existente
workbook = load_workbook('files/billing.xlsx')
pg_bill = workbook['Sheet1']

# Iterar sobre as linhas a partir da segunda linha
for row in pg_bill.iter_rows(min_row=2):
    row[1].value = '5547997459085'

# Salvar as alterações no arquivo Excel (apenas uma vez, após o loop)
workbook.save('files/billing.xlsx')

print('ok valores da coluna substituidos')