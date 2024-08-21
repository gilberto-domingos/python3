import os
import webbrowser
from datetime import datetime
from time import sleep
from urllib.parse import quote
from openpyxl import load_workbook
import pyautogui

workbook = load_workbook('files/billing.xlsx')
pg_bill = workbook['Sheet1']

date_now = datetime.now()

for row in pg_bill.iter_rows(min_row=2, max_row=22):
    company = row[0].value
    phone = row[1].value
    bill_str = row[2].value  # Leia a data como string
    bill = datetime.strptime(bill_str, '%d/%m/%Y')  # Converta para datetime

    if bill < date_now:
        payment = 'venceu'
    else:
        payment = 'vai vencer'

    msg = (f'Olá {company}, seu boleto {payment} no dia {bill.strftime("%d/%m/%Y")}. '
           f'Favor pagar no link https://www.seusistemapagamento.com')

    try:
        print('Opening the browser')
        print('Code running...')
        link_msg_whats = f'https://web.whatsapp.com/send?phone={phone}&text={quote(msg)}'
        webbrowser.open(link_msg_whats)
        sleep(15)
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        seta = pyautogui.locateCenterOnScreen('files/arrow.png')
        sleep(2)
        pyautogui.click(seta[0], seta[1])
        sleep(2)
        pyautogui.hotkey('ctrl', 'w')
        sleep(2)
        print('Cobrança enviada para : ', {company}, {phone})
    except Exception as e:
        print(f'Não foi possível enviar mensagem para {company}: {e}')
        with open('files/erros.csv', 'a', newline='', encoding='utf-8') as file:
            file.write(f'{company},{phone}{os.linesep}')
