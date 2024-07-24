import pandas as pd
from openpyxl import load_workbook

# Abrir o arquivo Excel existente
file_path = 'files/com_vendas_.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# Lendo os dados no DataFrame
df = pd.read_excel(file_path)

# Função para calcular a comissão
def calculate_commission(sale_value):
    sale_value = float(sale_value.replace('R$ ', '').replace(',', ''))
    commission = sale_value * 0.10
    return f'R$ {commission:.2f}'

# Adicionar a comissão calculada na coluna "Commission"
for index, row in df.iterrows():
    sale_value = row['Venda']
    commission = calculate_commission(sale_value)
    ws[f'H{index + 2}'] = commission  # A coluna "Commission" é a 8ª coluna (H)

# Salvar o novo arquivo Excel
output_file_path = 'files/calc_com_vendas_.xlsx'
wb.save(output_file_path)
print(f'File saved as {output_file_path}')
