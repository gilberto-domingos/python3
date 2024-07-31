import openpyxl
import pyperclip
import pyautogui
from time import sleep

workbook = openpyxl.load_workbook('bot_files/com_vendas.xlsx')

sheet_vendas = workbook['Sheet1']

for row in sheet_vendas.iter_rows(min_row=2):
    company = row[0].value
    pyperclip.copy(company)
    pyautogui.click(870, 325, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    cnpj = row[1].value
    pyperclip.copy(cnpj)
    pyautogui.click(869, 419, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    title = row[2].value
    pyperclip.copy(title)
    pyautogui.click(871, 513, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    process = row[3].value
    pyperclip.copy(process)
    pyautogui.click(870, 608, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    pyautogui.click(904, 665, duration=1)
    #sleep(1)
###
    product = row[4].value
    pyperclip.copy(product)
    pyautogui.click(873,325, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    name = row[5].value
    pyperclip.copy(name)
    pyautogui.click(874,417, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    weight = row[6].value
    pyperclip.copy(weight)
    pyautogui.click(874,513, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    size = row[7].value
    pyperclip.copy(size)
    pyautogui.click(875,605, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    pyautogui.click(904, 665, duration=1)
    #sleep(1)

    employee = row[8].value
    pyperclip.copy(employee)
    pyautogui.click(873,325, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    sale = row[9].value
    pyperclip.copy(sale)
    pyautogui.click(876,419, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    date = row[10].value
    pyperclip.copy(date)
    pyautogui.click(874,509, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    commission = row[11].value
    pyperclip.copy(commission)
    pyautogui.click(872,606, duration=1)
    pyautogui.hotkey('ctrl', 'v')

    pyautogui.click(904, 665, duration=1)

    pyautogui.click(904, 665, duration=1)

    print(f"Copied process: {process}")
    sleep(300)

'''
company = row[0].value
cnpj = row[1].value
title = row[2].value
process = row[3].value36636803

product = row[4].value
name = row[5].value
weight = row[6].value
size = row[7].value

employee = row[8].value
sale = row[9].value
date = row[10].value
commission = row[11].value
'''

'''
sheet = workbook.active

# Linha que deseja imprimir (começando em 1)
row_number = 2

# Iterar pelas linhas e imprimir a linha desejada
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    if row[0].row == row_number:
        for cell in row:
            print(cell.value, end=" ")
        print()






for col in sheet.iter_rows(min_row=2):
    print(col[1].value)


# Imprimir os valores das células
for row in sheet.iter_rows(values_only=True):
    print(row)

headers = [cell.value for cell in sheet[1]]
print('\t'.join(headers))

line_count = 0

for row in sheet.iter_rows(min_row=2, values_only=True):
    print('\t'.join(map(str, row)))
    line_count += 1
    if line_count == 3:
        break
'''
